#!/usr/bin/env python3
"""Import todos, recipes, schedule, and meals from Excel into Firebase."""

import openpyxl
import requests
import json
import re
import time

EXCEL_PATH = '/Users/user/Downloads/누나와 가계부.xlsx'
FIREBASE_BASE = 'https://our-diary-240f6-default-rtdb.firebaseio.com/ourdiary'
YEAR = 2026

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

def val(cell):
    v = cell.value
    if v is None:
        return ''
    return str(v).strip()

def put(path, data):
    url = f'{FIREBASE_BASE}/{path}.json'
    r = requests.put(url, json=data)
    if r.status_code not in (200, 201):
        print(f'  ERROR {r.status_code}: {path}')
    else:
        print(f'  PUT OK: {path}')
    return r

def post_fb(path, data):
    url = f'{FIREBASE_BASE}/{path}.json'
    r = requests.post(url, json=data)
    if r.status_code not in (200, 201):
        print(f'  ERROR {r.status_code}: {path}')
        return None
    else:
        key = r.json().get('name','?')
        print(f'  POST OK: {path} -> {key}')
        return key


# ──────────────────────────────────────────────────────────────────────────────
# 1. TODOS
# ──────────────────────────────────────────────────────────────────────────────
print('\n══════ 1. TODOS ══════')
ws = wb['해야할 일']

todo_items = []
for row in ws.iter_rows(min_row=1, max_row=20):
    for cell in row:
        v = val(cell)
        if v and cell.column == 2 and 'km' not in v.lower():
            # Skip the 쓰레기 note (multi-line with *)
            if '*' not in v and '쓰레기' not in v and '[' not in v:
                todo_items.append(v)

print(f'Todos to import: {todo_items}')

for item in todo_items:
    ts = int(time.time() * 1000)
    time.sleep(0.05)
    post_fb('todos', {
        'text': item,
        'done': False,
        'createdAt': ts
    })


# ──────────────────────────────────────────────────────────────────────────────
# 2. RECIPES
# ──────────────────────────────────────────────────────────────────────────────
print('\n══════ 2. RECIPES ══════')
ws_r = wb['레시피']

# Recipe 1: 닭도리탕 (rows 1-11)
title1 = ''
ingredients1_parts = []
steps1 = []

for row in ws_r.iter_rows(min_row=1, max_row=12):
    row_vals = {cell.column: val(cell) for cell in row if val(cell)}
    if not row_vals:
        continue

    col1 = row_vals.get(1, '')
    col2 = row_vals.get(2, '')

    if '닭도리탕' in col1:
        title1 = col1
    elif col1 == '재료':
        ingredients1_parts.append(f'재료: {col2}')
    elif col1 == '양념장':
        ingredients1_parts.append(f'양념장: {col2}')
    elif col1 == '순서' or col1.startswith('1.') or re.match(r'^\d+\.', col1):
        # first step row might have "순서" label
        if col1 == '순서' and col2:
            steps1.append(col2)
        elif re.match(r'^\d+\.', col1):
            steps1.append(col1)

# Steps are in col 1 rows 4-11
steps1 = []
for row in ws_r.iter_rows(min_row=4, max_row=12):
    row_vals = {cell.column: val(cell) for cell in row if val(cell)}
    col1 = row_vals.get(1, '')
    col2 = row_vals.get(2, '')
    if re.match(r'^\d+\.', col1):
        steps1.append(col1)
    elif col1 == '순서' and col2:
        steps1.append(col2)

recipe1 = {
    'title': title1,
    'ingredients': '\n'.join(ingredients1_parts),
    'steps': '\n'.join(steps1),
    'createdAt': int(time.time() * 1000)
}
print(f'Recipe 1: {title1}')
print(f'  ingredients: {recipe1["ingredients"][:80]}...')
print(f'  steps: {len(steps1)} steps')
post_fb('recipes', recipe1)

# Recipe 2: 두부된장찌개 (row 17)
time.sleep(0.1)
big_text = ''
for row in ws_r.iter_rows(min_row=13, max_row=20):
    row_vals = {cell.column: val(cell) for cell in row if val(cell)}
    col1 = row_vals.get(1, '')
    if col1 and len(col1) > 50:
        big_text = col1
        break

# Parse the big text into sections
title2 = '두부된장찌개'
ingredients2 = ''
steps2 = ''

if big_text:
    # Split on "2. 만드는 법" or similar
    parts = big_text.split('2. 만드는 법')
    if len(parts) == 2:
        ingredients_raw = parts[0].replace('1. 재료 준비\n\n', '').strip()
        steps_raw = parts[1].strip()
        ingredients2 = ingredients_raw
        steps2 = '2. 만드는 법\n\n' + steps_raw

recipe2 = {
    'title': title2,
    'ingredients': ingredients2 if ingredients2 else big_text[:500],
    'steps': steps2 if steps2 else '',
    'createdAt': int(time.time() * 1000)
}
print(f'Recipe 2: {title2}')
print(f'  ingredients: {recipe2["ingredients"][:80]}...')
post_fb('recipes', recipe2)


# ──────────────────────────────────────────────────────────────────────────────
# 3. SCHEDULE
# ──────────────────────────────────────────────────────────────────────────────
print('\n══════ 3. SCHEDULE ══════')
ws_s = wb['일정표']

# Read all rows with column info
all_rows = []
for row in ws_s.iter_rows(min_row=1):
    row_data = {}
    for cell in row:
        v = val(cell)
        if v:
            row_data[cell.column] = v
    all_rows.append(row_data)

# Parse month/day/event structure
# Structure: month header row, then week blocks of (day-row, event-row)
# Day rows: all values are numbers or "N 공휴일" format
# Event rows: values are text strings (not numbers)

def is_day_val(s):
    """Check if string represents a day number (possibly with holiday text)."""
    s = s.strip()
    # Pure float like "1.0"
    if re.match(r'^\d+\.0$', s):
        return True
    # "1 노동절" or "25 대체공휴일"
    if re.match(r'^\d+\s', s):
        return True
    # Day/week headers
    if s in ('일', '월', '화', '수', '목', '금', '토'):
        return True
    return False

def parse_day_num(s):
    """Extract day number from day value string."""
    m = re.match(r'^(\d+)', s)
    if m:
        return int(m.group(1))
    return None

# Identify month header rows
MONTH_NAMES = {'1월':1,'2월':2,'3월':3,'4월':4,'5월':5,'6월':6,
               '7월':7,'8월':8,'9월':9,'10월':10,'11월':11,'12월':12}

schedule_events = []
current_month = None
prev_day_row = None  # (row_index, col -> day_num mapping)

for i, row_data in enumerate(all_rows):
    if not row_data:
        continue

    values = list(row_data.values())

    # Check if month header
    for v in values:
        if v in MONTH_NAMES:
            current_month = MONTH_NAMES[v]
            prev_day_row = None
            break

    if current_month is None:
        continue

    # Skip header rows (일월화수목금토)
    if all(v in ('일','월','화','수','목','금','토') for v in values):
        continue

    # Check if this is a day row (most values are day-like)
    day_like = sum(1 for v in values if is_day_val(v))
    total = len(values)

    if total > 0 and day_like == total:
        # This is a day row - build col->day mapping
        col_to_day = {}
        for col, v in row_data.items():
            d = parse_day_num(v)
            if d:
                col_to_day[col] = d
        prev_day_row = col_to_day
    elif prev_day_row is not None:
        # This is an event row - match columns to days
        # Sort columns to find nearest day column
        day_cols = sorted(prev_day_row.keys())

        for col, event_text in sorted(row_data.items()):
            # Skip if it's a day-like value (shouldn't happen but safety check)
            if is_day_val(event_text):
                continue
            # Find closest day column at or before this event column
            matched_day = None
            for dc in day_cols:
                if dc <= col:
                    matched_day = prev_day_row[dc]
                else:
                    break

            if matched_day is None:
                # Try the first day column
                if day_cols:
                    matched_day = prev_day_row[day_cols[0]]

            if matched_day and event_text:
                # Split multi-line events
                lines = [l.strip() for l in event_text.split('\n') if l.strip()]
                for line in lines:
                    # Skip emoji-only or very short pure numeric
                    if re.match(r'^\d+$', line):
                        continue
                    schedule_events.append({
                        'month': current_month,
                        'day': matched_day,
                        'text': line
                    })
                    print(f'  {YEAR}-{current_month:02d}-{matched_day:02d}: {line}')

print(f'\nTotal schedule events: {len(schedule_events)}')

# Import to Firebase
imported_schedule = 0
for ev in schedule_events:
    m = ev['month']
    d = ev['day']
    month_key = f'{YEAR}-{m:02d}'
    day_key = str(d)
    key = post_fb(f'schedule/{month_key}/{day_key}', {
        'text': ev['text'],
        'date': f'{YEAR}-{m:02d}-{d:02d}',
        'createdAt': int(time.time() * 1000)
    })
    if key:
        imported_schedule += 1
    time.sleep(0.03)

print(f'Imported {imported_schedule} schedule events')


# ──────────────────────────────────────────────────────────────────────────────
# 4. MEALS (급식표)
# ──────────────────────────────────────────────────────────────────────────────
print('\n══════ 4. MEALS (급식표) ══════')
ws_m = wb['급식표']

all_meal_rows = []
for row in ws_m.iter_rows(min_row=1):
    row_data = {}
    for cell in row:
        v = val(cell)
        if v:
            row_data[cell.column] = v
    all_meal_rows.append(row_data)

meal_events = []
current_month_m = None
prev_day_row_m = None

for i, row_data in enumerate(all_meal_rows):
    if not row_data:
        continue

    values = list(row_data.values())

    # Check if month header
    for v in values:
        if v in MONTH_NAMES:
            new_month = MONTH_NAMES[v]
            if new_month != current_month_m:
                current_month_m = new_month
                prev_day_row_m = None
            break

    # First data rows before any month header -> April
    if current_month_m is None:
        # Check if this looks like day data
        day_like = sum(1 for v in values if is_day_val(v))
        if day_like > 0:
            current_month_m = 4  # Default to April

    if current_month_m is None:
        continue

    # Skip header rows
    if all(v in ('일','월','화','수','목','금','토','급식표') for v in values):
        continue
    if len(values) == 1 and list(values)[0] == '급식표':
        continue

    # Check if day row
    day_like = sum(1 for v in values if is_day_val(v))
    total = len(values)

    if total > 0 and day_like == total:
        col_to_day = {}
        for col, v in row_data.items():
            d = parse_day_num(v)
            if d:
                col_to_day[col] = d
        prev_day_row_m = col_to_day
    elif prev_day_row_m is not None:
        # Event/meal row
        day_cols = sorted(prev_day_row_m.keys())

        for col, meal_text in sorted(row_data.items()):
            if is_day_val(meal_text):
                continue

            # Find closest day column at or before
            matched_day = None
            for dc in day_cols:
                if dc <= col:
                    matched_day = prev_day_row_m[dc]
                else:
                    break

            if matched_day is None and day_cols:
                matched_day = prev_day_row_m[day_cols[0]]

            if matched_day and meal_text:
                # Filter out schedule-like texts (이모지 events)
                if '🎊' in meal_text:
                    continue

                meal_events.append({
                    'month': current_month_m,
                    'day': matched_day,
                    'text': meal_text
                })
                print(f'  {YEAR}-{current_month_m:02d}-{matched_day:02d}: {meal_text[:60]}')

print(f'\nTotal meal events: {len(meal_events)}')

# Import meals
imported_meals = 0
for ev in meal_events:
    m = ev['month']
    d = ev['day']
    month_key = f'{YEAR}-{m:02d}'
    day_key = str(d)

    # Meals: store as array of menu items (split by newline)
    lines = [l.strip() for l in ev['text'].split('\n') if l.strip()]
    menu_text = '\n'.join(lines)

    # Check if day already has meal data and merge
    url = f'{FIREBASE_BASE}/meals/{month_key}/{day_key}.json'
    existing = requests.get(url).json()

    if existing:
        # Append to existing
        if isinstance(existing, str):
            merged = existing + '\n' + menu_text
        else:
            merged = menu_text
        requests.put(url, json=merged)
        print(f'  MERGE: meals/{month_key}/{day_key}')
    else:
        requests.put(url, json=menu_text)
        print(f'  PUT: meals/{month_key}/{day_key}')

    imported_meals += 1
    time.sleep(0.03)

print(f'\nImported {imported_meals} meal entries')

print('\n══════ DONE ══════')
print(f'Todos: {len(todo_items)}')
print(f'Recipes: 2')
print(f'Schedule events: {imported_schedule}')
print(f'Meal entries: {imported_meals}')
